from django.contrib import admin
from django.utils.html import format_html
from django.http import HttpResponse
from .models import Pituto, Postulacion, Perfil


# ──────────────────────────────────────────────
# Acción personalizada: Exportar a Excel
# ──────────────────────────────────────────────
def exportar_pitutos_excel(modeladmin, request, queryset):
    """Exporta los pitutos seleccionados a un archivo Excel con formato."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        modeladmin.message_user(
            request,
            "❌ El módulo 'openpyxl' no está instalado. Ejecuta: pip install openpyxl",
            level='error'
        )
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pitutos"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    estado_fills = {
        'ACTIVO':    PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
        'OTORGADO':  PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid"),
        'FINALIZADO': PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid"),
    }

    # Encabezados
    headers = [
        "ID", "Título", "Oferente", "Email Oferente",
        "Cuidador Seleccionado", "Email Cuidador",
        "Estado", "Pago", "Tipo de Pago", "Categoría", "Comuna",
        "Calif. al Cuidador (1-5)", "Comentario al Cuidador",
        "Calif. al Oferente (1-5)", "Comentario al Oferente",
        "Fecha Creación", "Fecha Finalización",
    ]
    ws.append(headers)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    ws.row_dimensions[1].height = 30

    # Filas de datos
    for pituto in queryset.select_related('creador', 'cuidador_seleccionado'):
        cuidador = pituto.cuidador_seleccionado
        row = [
            pituto.id,
            pituto.titulo,
            pituto.creador.get_full_name() or pituto.creador.username,
            pituto.creador.email,
            (cuidador.get_full_name() or cuidador.username) if cuidador else "—",
            cuidador.email if cuidador else "—",
            pituto.get_estado_display(),
            pituto.pago,
            pituto.tipo_pago,
            pituto.get_categoria_display(),
            pituto.comuna or "Remoto",
            pituto.calificacion_a_cuidador or "—",
            pituto.comentario_a_cuidador or "—",
            pituto.calificacion_a_oferente or "—",
            pituto.comentario_a_oferente or "—",
            pituto.fecha_creacion.strftime("%d/%m/%Y %H:%M") if pituto.fecha_creacion else "—",
            pituto.fecha_finalizacion.strftime("%d/%m/%Y %H:%M") if pituto.fecha_finalizacion else "—",
        ]
        ws.append(row)

        # Colorear fila según estado
        row_idx = ws.max_row
        fill = estado_fills.get(pituto.estado)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if fill:
                cell.fill = fill

    # Ajustar anchos de columna automáticamente
    col_widths = [6, 35, 25, 30, 25, 30, 15, 14, 14, 18, 18, 8, 40, 8, 40, 20, 20]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="pitutea_pitutos.xlsx"'
    wb.save(response)
    return response

exportar_pitutos_excel.short_description = "📥 Exportar seleccionados a Excel"


# ──────────────────────────────────────────────
# Admin: Perfil
# ──────────────────────────────────────────────
@admin.register(Perfil)
class PerfilAdmin(admin.ModelAdmin):
    list_display = ('usuario', 'rol', 'rut', 'estado_verificacion', 'carnet_cuidador')
    list_filter = ('rol', 'estado_verificacion')
    search_fields = ('usuario__username', 'usuario__first_name', 'usuario__last_name', 'rut')

    readonly_fields = ('ver_carnet_cuidador',)

    fieldsets = (
        ('Información Básica', {
            'fields': ('usuario', 'rol', 'rut', 'telefono', 'direccion', 'comuna')
        }),
        ('Información Cuidador', {
            'fields': ('categoria_interes', 'habilidades', 'carnet_cuidador', 'ver_carnet_cuidador'),
        }),
        ('Verificación y Estado', {
            'fields': ('estado_verificacion', 'observaciones_verificacion'),
        }),
    )

    def ver_carnet_cuidador(self, obj):
        if obj.carnet_cuidador:
            return format_html(
                '<div class="carnet-preview-container" style="margin-top: 10px;">'
                '  <a href="{0}" target="_blank" style="display: inline-block; border: 2px solid #eaeaea; border-radius: 8px; overflow: hidden; box-shadow: 0 4px 12px rgba(0,0,0,0.08); transition: all 0.3s ease;">'
                '    <img src="{0}" alt="Carnet Cuidador" class="carnet-preview-img" style="max-width: 450px; max-height: 300px; display: block; object-fit: contain;" />'
                '  </a>'
                '  <p class="help carnet-preview-help" style="margin: 8px 0 0 0; font-size: 0.85em; color: #666;">'
                '    Haz clic en la imagen para verla en pantalla completa'
                '  </p>'
                '</div>',
                obj.carnet_cuidador.url
            )
        return "No se ha cargado una imagen de carnet."
    ver_carnet_cuidador.short_description = "Vista Previa del Carnet"


# ──────────────────────────────────────────────
# Admin: Pituto
# ──────────────────────────────────────────────
def stars_display(value):
    """Helper: converts numeric rating to star string."""
    if value is None:
        return "—"
    return "⭐" * value + "☆" * (5 - value) + f" ({value}/5)"


@admin.register(Pituto)
class PitutoAdmin(admin.ModelAdmin):
    list_display = (
        'titulo', 'creador', 'estado_badge', 'cuidador_seleccionado',
        'calificacion_a_cuidador', 'calificacion_a_oferente',
        'pago', 'tipo_pago', 'fecha_creacion', 'fecha_finalizacion'
    )
    list_filter = ('estado', 'activo', 'tipo_pago', 'categoria')
    search_fields = ('titulo', 'descripcion', 'creador__username', 'cuidador_seleccionado__username')
    readonly_fields = ('fecha_creacion', 'fecha_finalizacion', 'calificaciones_resumen')
    actions = [exportar_pitutos_excel]

    fieldsets = (
        ('Información del Pituto', {
            'fields': ('titulo', 'descripcion', 'categoria', 'comuna', 'pago', 'pago_anterior', 'tipo_pago', 'flexibilidad')
        }),
        ('Estado y Ciclo de Vida', {
            'fields': ('activo', 'estado', 'cuidador_seleccionado', 'fecha_creacion', 'fecha_finalizacion')
        }),
        ('Calificaciones', {
            'fields': ('calificaciones_resumen', 'calificacion_a_cuidador', 'comentario_a_cuidador', 'calificacion_a_oferente', 'comentario_a_oferente'),
        }),
    )

    def estado_badge(self, obj):
        colors = {
            'ACTIVO':    ('28a745', 'Activo'),
            'OTORGADO':  ('ffc107', 'Trabajo Otorgado'),
            'FINALIZADO': ('0d6efd', 'Finalizado'),
        }
        color, label = colors.get(obj.estado, ('6c757d', obj.estado))
        return format_html(
            '<span style="background:#{};color:{};padding:3px 10px;border-radius:20px;font-size:0.78em;font-weight:600;">{}</span>',
            color,
            'white' if obj.estado != 'OTORGADO' else '#333',
            label
        )
    estado_badge.short_description = "Estado"

    def calificaciones_resumen(self, obj):
        lines = []
        if obj.calificacion_a_cuidador:
            lines.append(f"⭐ Oferente → Cuidador: {stars_display(obj.calificacion_a_cuidador)}")
            if obj.comentario_a_cuidador:
                lines.append(f'   "{obj.comentario_a_cuidador}"')
        if obj.calificacion_a_oferente:
            lines.append(f"⭐ Cuidador → Oferente: {stars_display(obj.calificacion_a_oferente)}")
            if obj.comentario_a_oferente:
                lines.append(f'   "{obj.comentario_a_oferente}"')
        if not lines:
            return "Sin calificaciones aún."
        return format_html("<pre style='margin:0;font-size:0.9em;'>{}</pre>", "\n".join(lines))
    calificaciones_resumen.short_description = "Resumen de Calificaciones"


# ──────────────────────────────────────────────
# Admin: Postulacion
# ──────────────────────────────────────────────
@admin.register(Postulacion)
class PostulacionAdmin(admin.ModelAdmin):
    list_display = ('pituto', 'usuario', 'fecha_postulacion')
    list_filter = ('fecha_postulacion',)
